"""
exporter.py
===========
Módulo de exportación de resultados a Excel (.xlsx) y CSV.
Genera archivos con formato profesional, incluyendo:
  - Hoja de resultados detallados
  - Hoja de resumen por sitio
  - Formato con colores, anchos de columna y filtros automáticos
"""

import csv
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict

logger = logging.getLogger("telecom_searcher.exporter")


def _get_output_path(query: str, output_dir: str, extension: str) -> Path:
    """
    Genera la ruta de salida con timestamp y nombre de query saneado.
    
    Args:
        query: Query original de búsqueda.
        output_dir: Directorio de salida.
        extension: Extensión del archivo ('xlsx' o 'csv').

    Returns:
        Path completo del archivo de salida.
    """
    # Sanear el nombre de la query para usar como nombre de archivo
    safe_query = "".join(c if c.isalnum() or c in " _-" else "_" for c in query)
    safe_query = safe_query[:50].strip().replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"telecom_search_{safe_query}_{timestamp}.{extension}"

    os.makedirs(output_dir, exist_ok=True)
    return Path(output_dir) / filename


def export_to_csv(results: List[Dict], query: str, output_dir: str = "output") -> str:
    """
    Exporta resultados a un archivo CSV.

    Args:
        results: Lista de dicts con resultados de búsqueda.
        query: Query original.
        output_dir: Directorio de salida.

    Returns:
        Ruta del archivo creado.
    """
    if not results:
        logger.warning("No hay resultados para exportar a CSV.")
        return ""

    path = _get_output_path(query, output_dir, "csv")

    fieldnames = [
        "site_id", "site_name", "query", "language",
        "url", "link_text", "timestamp"
    ]

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

    logger.info(f"CSV exportado: {path} ({len(results)} registros)")
    return str(path)


def export_to_excel(
    results: List[Dict],
    query: str,
    output_dir: str = "output",
    sites: List[Dict] = None,
) -> str:
    """
    Exporta resultados a un archivo Excel con formato profesional.
    Incluye dos hojas: 'Resultados' y 'Resumen'.

    Args:
        results: Lista de dicts con resultados de búsqueda.
        query: Query original.
        output_dir: Directorio de salida.
        sites: Lista de sitios configurados (para el resumen).

    Returns:
        Ruta del archivo creado.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl no instalado. Ejecuta: pip install openpyxl")
        return export_to_csv(results, query, output_dir)

    if not results:
        logger.warning("No hay resultados para exportar a Excel.")
        return ""

    path = _get_output_path(query, output_dir, "xlsx")

    wb = openpyxl.Workbook()

    # ─── Estilos ────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F4E79")    # Azul oscuro
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="D6E4F0")    # Azul claro
    link_font   = Font(color="1155CC", underline="single")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ─── Hoja 1: Resultados ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resultados"

    # Metadata en las primeras filas
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Búsqueda Telecom: \"{query}\""
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = center_align

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total: {len(results)} links"
    ws["A2"].font = Font(size=10, italic=True, color="666666")
    ws["A2"].alignment = center_align

    ws.append([])  # Fila vacía

    # Encabezados
    headers = ["Organismo", "Nombre Completo", "Query Usada", "Idioma", "URL", "Texto del Link", "Fecha/Hora"]
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Filas de datos
    for row_idx, result in enumerate(results, header_row + 1):
        data = [
            result.get("site_id", ""),
            result.get("site_name", ""),
            result.get("query", ""),
            result.get("language", ""),
            result.get("url", ""),
            result.get("link_text", ""),
            result.get("timestamp", ""),
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = left_align

            # Alternar colores de fila
            if row_idx % 2 == 0:
                cell.fill = alt_fill

            # Formato especial para URLs (columna 5)
            if col_idx == 5 and value:
                cell.font = link_font
                cell.hyperlink = value

    # Anchos de columna optimizados
    col_widths = [12, 40, 35, 10, 60, 45, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Activar filtros automáticos
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(results)}"

    # Congelar encabezados
    ws.freeze_panes = f"A{header_row + 1}"

    # ─── Hoja 2: Resumen ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Resumen por Organismo"
    ws2["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws2["A1"].alignment = center_align

    ws2.append([])

    sum_headers = ["Organismo (ID)", "Nombre", "Links Encontrados", "Idioma del Sitio"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Contar links por sitio
    counts: Dict[str, int] = {}
    for r in results:
        counts[r["site_id"]] = counts.get(r["site_id"], 0) + 1

    # Mapear site_id → nombre e idioma usando la config de sitios
    site_info = {}
    if sites:
        for s in sites:
            site_info[s["id"]] = {
                "name": s["name"],
                "language": s.get("language", "en"),
            }

    # Llenar resumen con todos los sitios (incluso los que no encontraron nada)
    all_site_ids = list(site_info.keys()) if site_info else list(counts.keys())

    for row_idx, site_id in enumerate(all_site_ids, 4):
        info = site_info.get(site_id, {})
        count = counts.get(site_id, 0)

        row_data = [
            site_id,
            info.get("name", site_id),
            count,
            info.get("language", "—"),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center_align if col_idx != 2 else left_align

            if row_idx % 2 == 0:
                cell.fill = alt_fill

            # Resaltar en verde si encontró links, rojo si no
            if col_idx == 3:
                cell.font = Font(
                    bold=True,
                    color="1E7B34" if count > 0 else "C00000"
                )

    # Fila de total
    total_row = 4 + len(all_site_ids)
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=3, value=len(results)).font = Font(bold=True)

    # Anchos de columna del resumen
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 15

    # ─── Guardar ─────────────────────────────────────────────────────────────
    wb.save(path)
    logger.info(f"Excel exportado: {path} ({len(results)} registros, 2 hojas)")
    return str(path)


def print_summary(results: List[Dict], sites: List[Dict] = None):
    """
    Imprime un resumen de resultados en la consola.
    """
    print("\n" + "=" * 60)
    print(f"  RESUMEN DE RESULTADOS")
    print("=" * 60)
    print(f"  Total de links encontrados: {len(results)}")
    print()

    # Contar por sitio
    counts: Dict[str, int] = {}
    for r in results:
        counts[r["site_id"]] = counts.get(r["site_id"], 0) + 1

    print(f"  {'Organismo':<12} {'Links':>6}  Status")
    print(f"  {'-'*35}")

    all_ids = [s["id"] for s in sites] if sites else list(counts.keys())
    for site_id in all_ids:
        count = counts.get(site_id, 0)
        status = "✓" if count > 0 else "✗ Sin resultados"
        print(f"  {site_id:<12} {count:>6}  {status}")

    print("=" * 60 + "\n")
